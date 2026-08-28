"""
Matchday Insights — record/streak precompute script.

Reads the historical rankings workbook (Historical Rank + Historical
Scores sheets) and, for every club, computes the fields needed to drive
the "record" badges and #1 streak line on the main rankings table:

  rank_ath, rank_atl        bool  — currently at all-time best/worst rank
  rank_high, rank_low       int   — the all-time best/worst rank values
  rank_streak               int   — consecutive matchdays currently held at rank_high
                                    (only meaningful when rank_ath is True)
  rank_streak_since         str   — date (M/D/YYYY) the CURRENT streak began
  rank_prev_since           str   — date of the most recent PRIOR time the club
                                    held rank_high, before this streak (None if
                                    this is the first time ever)
  elo_ath, elo_atl          bool  — currently at all-time best/worst rating
  top1_streak                int  — consecutive matchdays at rank #1 (0 if not #1)

Output: a dict keyed by club name -> field dict, written to records.json.
Merge this into your CLUBS array generation step by club name.

Usage:
    python3 compute_records.py New_Historical_Rankings_Revamp.xlsx records.json
"""
import sys
import json
import openpyxl

RANK_SHEET = "Historical Rank"
SCORE_SHEET = "Historical Scores"

RANK_META_COLS = 14  # Club, Nation, #1, Top5, Top10, Top20, Top50, Top100, High, High Count, Most Recent, Low, Low Count, Most Recent2
SCORE_META_COLS = 4  # Club, Nation, High, Low


def consecutive_streak(values, target):
    """Count consecutive matches of `target` at the start of `values`
    (values is most-recent-first). Returns (streak_len, date_idx_after_streak)."""
    n = 0
    for v in values:
        if v == target:
            n += 1
        else:
            break
    return n


def find_prev_occurrence(values, target, skip):
    """After skipping the first `skip` (current streak) entries, find the
    index of the next most-recent entry equal to target. Returns index or None."""
    for i in range(skip, len(values)):
        if values[i] == target:
            return i
    return None


def load_rank_sheet(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[RANK_SHEET]
    rows = ws.iter_rows(min_row=2, values_only=True)
    out = {}
    for row in rows:
        if row[0] is None:
            continue
        club = row[0]
        high = row[8]
        high_count = row[9]
        low = row[11]
        low_count = row[12]
        rank_series = list(row[RANK_META_COLS:])
        # trim trailing Nones (short rows)
        while rank_series and rank_series[-1] is None:
            rank_series.pop()
        out[club] = {
            "rank_high": high,
            "rank_high_count": high_count,
            "rank_low": low,
            "rank_low_count": low_count,
            "rank_series": rank_series,  # most-recent-first
        }
    return out


def load_score_sheet(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SCORE_SHEET]
    rows = ws.iter_rows(min_row=2, values_only=True)
    out = {}
    for row in rows:
        if row[0] is None:
            continue
        club = row[0]
        high = row[2]
        low = row[3]
        elo_series = list(row[SCORE_META_COLS:])
        while elo_series and elo_series[-1] is None:
            elo_series.pop()
        out[club] = {
            "elo_high": high,
            "elo_low": low,
            "elo_series": elo_series,
        }
    return out


def load_dates(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[RANK_SHEET]
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    date_cols = list(row1[RANK_META_COLS:])
    return date_cols


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 compute_records.py <workbook.xlsx> <output.json>")
        sys.exit(1)
    xlsx_path, out_path = sys.argv[1], sys.argv[2]

    dates = load_dates(xlsx_path)
    rank_data = load_rank_sheet(xlsx_path)
    score_data = load_score_sheet(xlsx_path)

    result = {}
    for club, rd in rank_data.items():
        series = rd["rank_series"]
        if not series:
            continue
        current_rank = series[0]
        high = rd["rank_high"]
        low = rd["rank_low"]

        rank_ath = (current_rank == high)
        rank_atl = (current_rank == low)

        rank_streak = 0
        rank_streak_since = None
        rank_prev_since = None
        if rank_ath:
            rank_streak = consecutive_streak(series, high)
            since_idx = rank_streak - 1
            if since_idx < len(dates):
                rank_streak_since = dates[since_idx]
            prev_idx = find_prev_occurrence(series, high, rank_streak)
            if prev_idx is not None and prev_idx < len(dates):
                rank_prev_since = dates[prev_idx]

        top1_streak = 0
        if current_rank == 1:
            top1_streak = consecutive_streak(series, 1)

        sd = score_data.get(club, {})
        elo_series = sd.get("elo_series", [])
        elo_high = sd.get("elo_high")
        elo_low = sd.get("elo_low")
        current_elo = elo_series[0] if elo_series else None
        elo_ath = (current_elo is not None and elo_high is not None and current_elo >= elo_high - 1e-6)
        elo_atl = (current_elo is not None and elo_low is not None and current_elo <= elo_low + 1e-6)

        result[club] = {
            "rank_ath": rank_ath,
            "rank_atl": rank_atl,
            "rank_high": high,
            "rank_low": low,
            "rank_streak": rank_streak,
            "rank_streak_since": rank_streak_since,
            "rank_prev_since": rank_prev_since,
            "elo_ath": elo_ath,
            "elo_atl": elo_atl,
            "top1_streak": top1_streak,
        }

    with open(out_path, "w") as f:
        json.dump(result, f, indent=1)
    print(f"Wrote {len(result)} clubs to {out_path}")


if __name__ == "__main__":
    main()
